import csv

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import DeleteView, TemplateView
from django.views.generic import ListView, View
from django.views.generic import UpdateView

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from rate import model_choices as mch
from rate.models import Rate
from rate.utils import display

from src.rate.model_choices import CURRENCY_TYPE_CHOICES, RATE_TYPE_CHOICES, SOURCE_CHOICES


class RateList(ListView):
    queryset = Rate.objects.all()
    template_name = 'rate-list.html'

    def get_source_display(self, source):
        return SOURCE_CHOICES[source]

    def get_currency_type_display(self, currency_type):
        return CURRENCY_TYPE_CHOICES[currency_type]

    def get_type_display(self, type_):
        return RATE_TYPE_CHOICES[type_]


class LatestRatesView(TemplateView):
    template_name = 'latest-rates.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        objects_list = []
        for source in mch.SOURCE_CHOICES:
            source = source[0]
            for currency_type in mch.CURRENCY_TYPE_CHOICES:
                currency_type = currency_type[0]
                for type_ in mch.RATE_TYPE_CHOICES:
                    type_ = type_[0]
                    rate = Rate.objects.filter(
                        source=source,
                        type=type_,
                        currency_type=currency_type).last()
                    if rate is not None:
                        objects_list.append(rate)

        context['object_list'] = objects_list
        return context


class RateDelete(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    queryset = Rate.objects.all()
    template_name = 'rate_delete.html'
    success_url = reverse_lazy('rate:list')

    def test_func(self):
        return self.request.user.is_superuser

    def get_object(self, queryset=None):
        pk = self.kwargs.get(self.pk_url_kwarg)
        if pk is not None:
            obj = self.get_queryset().get(id=pk)
        return obj


class RateEdit(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    queryset = Rate.objects.all()
    template_name = 'rate_edit.html'
    success_url = reverse_lazy('rate:list')
    fields = ('source', 'currency_type', 'type', 'amount')

    def test_func(self):
        return self.request.user.is_superuser

    def get_object(self, queryset=None):
        pk = self.kwargs.get(self.pk_url_kwarg)
        if pk is not None:
            obj = self.get_queryset().get(id=pk)
        return obj


class RateDownloadCSV(View):
    HEADERS = [
        'id',
        'source',
        'created',
        'type',
        'amount',
    ]
    queryset = Rate.objects.all().iterator()

    def get(self, request):
        response = self.get_response()

        writer = csv.writer(response)
        writer.writerow(self.__class__.HEADERS)

        for rate in self.queryset:
            values = []
            for attr in self.__class__.HEADERS:
                values.append(display(rate, attr))

            writer.writerow(values)
        return response

    def get_response(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rates.csv"'
        return response


class RateDownloadXLSX(View):
    HEADERS = [
        'id',
        'source',
        'created',
        'type',
        'amount',
    ]
    queryset = Rate.objects.all().iterator()

    def get(self, request):
        workbook = Workbook()
        sheet_name = workbook.sheetnames
        if sheet_name:
            sheet = workbook.get_sheet_by_name(sheet_name[0])
            for i, item in enumerate(self.HEADERS):
                cell = sheet.cell(row=1, column=i + 1)
                cell.value = self.HEADERS[i]
        row = 1
        for rate in self.queryset:
            row += 1
            for i, item in enumerate(self.HEADERS):
                cell = sheet.cell(row=row, column=i + 1)
                cell.value = display(rate, self.HEADERS[i])

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Rates.xlsx"'

        return response
